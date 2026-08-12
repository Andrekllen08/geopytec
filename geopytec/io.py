import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from io import StringIO
import os

def ler(conteudo_ou_caminho):# formata a tabela de dados
    """
    Lê o arquivo do equipamento, extrai metadados e separa os DataFrames
    de Adensamento e Cisalhamento.
    """
    if isinstance(conteudo_ou_caminho, str):
        with open(conteudo_ou_caminho, 'r', encoding='utf-8', errors='ignore') as f:
            texto = f.read()
    else:
        texto = conteudo_ou_caminho.decode('utf-8', errors='ignore')

    texto_limpo = texto.replace('\r\n', '\n').replace('\r', '\n')
    secoes = texto_limpo.split('\n\n')

    metadados = {}
    df_adensamento = None
    df_cisalhamento = None

    for sec in secoes:
        linhas = [l.strip() for l in sec.strip().split('\n') if l.strip()]
        if not linhas:
            continue

        titulo_secao = linhas[0].replace('\ufeff', '').replace(':', '').strip().lower()

        # 1. Metadados
        if any(chave in titulo_secao for chave in ['amostra', 'corpo de prova', 'metadados']):
            if len(linhas) >= 2:
                colunas = [c.strip() for c in linhas[1].split(';')]
                valores = [v.strip() for v in linhas[2].split(';')] if len(linhas) > 2 else []
                for c, v in zip(colunas, valores):
                    if c:
                        metadados[c] = v

        # 2. Adensamento
        elif 'adensamento' in titulo_secao:
            dados_csv = '\n'.join(linhas[1:])
            df_adensamento = pd.read_csv(StringIO(dados_csv), sep=';', decimal=',')

        # 3. Cisalhamento
        elif 'cisalhamento' in titulo_secao:
            dados_csv = '\n'.join(linhas[1:])
            df_cisalhamento = pd.read_csv(StringIO(dados_csv), sep=';', decimal=',')

    return metadados, df_adensamento, df_cisalhamento


def upload(caminho_ou_buffer, largura_caixa_mm=100.0):# Lê os dados
    """
    Carrega o arquivo (caminho ou buffer do Streamlit) e devolve
    o dicionário estruturado para os módulos da geopytec.
    """
    meta, df_ad, df_cis = ler(caminho_ou_buffer)

    if df_cis is not None:
        df_cis['Deformacao_%'] = (df_cis['Deslocamento Horizontal (mm)'] / largura_caixa_mm) * 100.0

    return {
        'metadados': meta,
        'adensamento': df_ad,
        'cisalhamento': df_cis,
        'resumo_tecnico': {
            'largura_caixa_mm': largura_caixa_mm,
            'tensao_normal_efetiva_kPa': float(meta.get('Tensao Normal', 0)) if meta else 0
        }
    }


    # ====================================================================
# FUNÇÃO: EXPORTAR RELATÓRIO EXCEL
# ====================================================================
def exportar_relatorio(ensaio_processado, c_kpa, phi_graus, arquivo_saida='relatorio_cisalhamento.xlsx'):
    """Gera um relatório profissional em Excel com os resultados dos ensaios."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Ensaios"

    # 1. Configurando Cores e Estilos Profissionais
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid") # Cinza chumbo
    zebra_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    zebra_fill_2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") # Cinza bem claro
    
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Side(style='thin', color='D3D3D3')
    border_style = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    align_center = Alignment(horizontal='center', vertical='center')

    # 2. Título Geral
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "RELATÓRIO DE ENSAIO DE CISALHAMENTO DIRETO"
    title_cell.font = Font(size=14, bold=True, color="2F4F4F")
    title_cell.alignment = align_center

    # 3. Cabeçalhos da Tabela
    headers = ["Amostra", "Tensão Normal (kPa)", "Tensão Ruptura (kPa)", "Deformação Ruptura (%)", "Critério Aplicado"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_style

    # 4. Escrevendo os Dados (com efeito Zebra)
    row = 4
    for idx, (nome, dados) in enumerate(ensaio_processado.items()):
        res = dados.get('resumo_tecnico', {})
        fill_color = zebra_fill_1 if idx % 2 == 0 else zebra_fill_2
        
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=res.get('tensao_normal_efetiva_kPa', 0))
        ws.cell(row=row, column=3, value=res.get('tau_ruptura_kPa', 0))
        ws.cell(row=row, column=4, value=res.get('deformacao_ruptura_%', 0))
        ws.cell(row=row, column=5, value=res.get('criterio_aplicado', '-'))
        
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.alignment = align_center
            cell.border = border_style
            cell.fill = fill_color
        row += 1

    # 5. Adicionando Parâmetros de Mohr-Coulomb
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    section_header = ws[f'A{row}']
    section_header.value = "PARÂMETROS DE RESISTÊNCIA (MOHR-COULOMB)"
    section_header.font = Font(bold=True, color="FFFFFF")
    section_header.fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    section_header.alignment = align_center
    
    row += 1
    ws[f'A{row}'] = "Coesão Efetiva (c')"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"{c_kpa:.2f}"
    ws[f'C{row}'] = "kPa"
    
    row += 1
    ws[f'A{row}'] = "Ângulo de Atrito Efetivo (φ')"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"{phi_graus:.2f}"
    ws[f'C{row}'] = "Graus"
    
    for r in range(row-1, row+1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border_style
            if c > 1:
                ws.cell(row=r, column=c).alignment = align_center

    # 6. Ajuste Automático de Largura das Colunas
    larguras = [25, 25, 25, 28, 22]
    for col, width in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Salva e finaliza
    wb.save(arquivo_saida)
    print(f"Relatório exportado com sucesso para: {arquivo_saida}")
    return arquivo_saida
